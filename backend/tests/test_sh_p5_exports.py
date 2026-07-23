"""SH-P5-5/6: export an approved PVC run to Excel / PDF.

Like `test_sh_p5_bills_get.py`, the route SQL uses Postgres-specific `::text`
casts that aiosqlite can't parse, so we stub `session.execute` at the
boundary and drive the handlers directly. Two gates carry the security
contract and so get dedicated tests: wrong-tenant → 404 (indistinguishable
from "doesn't exist") and not-Approved → 422 `run_not_approved`. The happy
path asserts the response is a real, non-empty file with the right media
type and an attachment disposition.
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock

import pytest

from api.exports import export_run_excel, export_run_pdf
from services.auth import AuthUser
from services.errors import NotFoundProblem, RunNotApprovedProblem
from services.exports import build_run_excel, build_run_pdf


def _user() -> AuthUser:
    return AuthUser(
        user_id="user-1",
        tenant_id="tenant-A",
        auth_id="auth-1",
        email="t@example.com",
        display_name="t@example.com",
    )


def _session_with(*results: tuple[str, object]) -> AsyncMock:
    """AsyncSession stub: each (kind, payload) feeds one `session.execute()`.
    kind="first" → `.mappings().first()`; kind="all" → `.mappings().all()`."""
    session = AsyncMock()
    mocked = []
    for kind, payload in results:
        result = MagicMock()
        mappings = MagicMock()
        if kind == "first":
            mappings.first.return_value = payload
        elif kind == "all":
            mappings.all.return_value = payload
        else:  # pragma: no cover — defensive
            raise ValueError(f"unknown result kind: {kind}")
        result.mappings.return_value = mappings
        mocked.append(result)
    session.execute = AsyncMock(side_effect=mocked)
    return session


_APPROVED_RUN = {
    "id": "run-1",
    "contract_id": "contract-1",
    "status": "Approved",
    "approved_by": "alice@example.com",
    "approved_at": datetime(2026, 6, 1, 10, 0, 0),
    "created_at": datetime(2026, 5, 31, 9, 0, 0),
    "tender_number": "T-123",
    "work_description": "Bridge renewal",
    "loa_number": "LOA-42",
    "loa_date": date(2025, 1, 15),
    "contractor_name": "Acme Infra",
    "base_month": date(2025, 1, 1),
    "railway_zone": "WR",
    "quarter_used": "Q2",
    "w_derivation": {
        "on_account_amount": "500000.00",
        "cement": "20000.00",
        "steel_angles": "10000.00",
        "steel_plates": "5000.00",
        "steel_tmt": "15000.00",
        "steel_other": "2500.00",
        "technical_withheld": "7500.00",
        "recoveries_affecting_pvc": "10000.00",
        "extra_items": "5000.00",
        "w": "425000.00",
    },
    "bill_snapshot": {"prior_negative_carry_forward": "1200.00"},
}

_COMPONENTS = [
    {
        "category": "cement",
        "eligible_amount": Decimal("100000.00"),
        "base_index": Decimal("130.2"),
        "current_avg_index": Decimal("133.5"),
        "weight": Decimal("0.10"),
        "pvc_value": Decimal("253.45"),
    },
    {
        "category": "steel",
        "eligible_amount": Decimal("200000.00"),
        "base_index": Decimal("57812.5"),
        "current_avg_index": Decimal("61917.5"),
        "weight": Decimal("0.25"),
        "pvc_value": Decimal("3548.10"),
    },
]

_APPROVED_RUNS = [
    {
        "id": "run-1",
        "bill_number": "RA-2",
        "quarter_used": "Q2",
        "w_amount": "425000.00",
        "total_pvc": Decimal("3801.55"),
    },
    {
        "id": "run-0",
        "bill_number": "RA-1",
        "quarter_used": "Q1",
        "w_amount": "300000.00",
        "total_pvc": Decimal("2500.00"),
    },
]


# ── Gating ──────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_excel_wrong_tenant_is_404():
    session = _session_with(("first", None))  # run not visible to tenant
    with pytest.raises(NotFoundProblem):
        await export_run_excel("run-1", user=_user(), session=session)


@pytest.mark.asyncio
async def test_export_pdf_wrong_tenant_is_404():
    session = _session_with(("first", None))
    with pytest.raises(NotFoundProblem):
        await export_run_pdf("run-1", user=_user(), session=session)


@pytest.mark.asyncio
async def test_export_excel_unapproved_is_422():
    draft = {**_APPROVED_RUN, "status": "Draft"}
    session = _session_with(("first", draft))
    with pytest.raises(RunNotApprovedProblem) as exc:
        await export_run_excel("run-1", user=_user(), session=session)
    assert exc.value.code == "run_not_approved"
    assert exc.value.status_code == 422


@pytest.mark.asyncio
async def test_export_pdf_unapproved_is_422():
    session = _session_with(("first", {**_APPROVED_RUN, "status": "Superseded"}))
    with pytest.raises(RunNotApprovedProblem):
        await export_run_pdf("run-1", user=_user(), session=session)


# ── Happy path ───────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_excel_returns_xlsx_attachment():
    session = _session_with(
        ("first", _APPROVED_RUN),
        ("all", _COMPONENTS),
        ("all", _APPROVED_RUNS),
    )
    resp = await export_run_excel("run-1", user=_user(), session=session)
    assert resp.status_code == 200
    assert resp.media_type.endswith("spreadsheetml.sheet")
    assert resp.headers["Content-Disposition"] == 'attachment; filename="pvc_run_run-1.xlsx"'
    # .xlsx is a zip container — magic bytes "PK".
    assert resp.body[:2] == b"PK"
    assert session.execute.await_count == 3
    wb = _load_workbook(resp.body)
    assert wb.sheetnames == ["Cover", "Bill", "W Derivation"]
    assert wb["Cover"]["B3"].value == "T-123 — Bridge renewal"
    assert wb["Cover"]["C12"].value == 425000
    assert wb["W Derivation"]["B13"].value == 425000
    primary_query = str(session.execute.await_args_list[0].args[0])
    for projection in (
        "r.w_derivation",
        "r.bill_snapshot",
        "c.work_description",
        "c.loa_number",
        "c.base_month",
        "c.railway_zone",
    ):
        assert projection in primary_query
    sibling_query = session.execute.await_args_list[2]
    assert "r.status = 'Approved'" in str(sibling_query.args[0])
    assert "COALESCE(" in str(sibling_query.args[0])
    assert "SUM(pc.pvc_value)" in str(sibling_query.args[0])
    assert sibling_query.args[1] == {"cid": "contract-1"}


@pytest.mark.asyncio
async def test_export_pdf_returns_pdf_attachment():
    session = _session_with(("first", _APPROVED_RUN), ("all", _COMPONENTS))
    resp = await export_run_pdf("run-1", user=_user(), session=session)
    assert resp.status_code == 200
    assert resp.media_type == "application/pdf"
    assert resp.headers["Content-Disposition"] == 'attachment; filename="pvc_run_run-1.pdf"'
    assert resp.body[:4] == b"%PDF"
    assert "w_derivation" not in str(session.execute.await_args_list[0].args[0])


# ── Pure generators ──────────────────────────────────────────────────────────


def test_build_run_excel_is_nonempty_zip():
    out = build_run_excel(_APPROVED_RUN, _COMPONENTS)
    assert isinstance(out, bytes) and out[:2] == b"PK" and len(out) > 0


def test_build_run_pdf_is_nonempty_pdf():
    out = build_run_pdf(_APPROVED_RUN, _COMPONENTS)
    assert isinstance(out, bytes) and out[:4] == b"%PDF" and len(out) > 0


def test_build_run_excel_handles_empty_components():
    out = build_run_excel(_APPROVED_RUN, [])
    assert out[:2] == b"PK"


# ── P8-REVIEW parity: submission-sheet column order, formats, live total ────


def _load_workbook(out: bytes):
    from io import BytesIO

    from openpyxl import load_workbook

    return load_workbook(BytesIO(out))


def _load_sheet(out: bytes):
    return _load_workbook(out)["Bill"]


def test_build_run_excel_submission_column_order():
    ws = _load_sheet(build_run_excel(_APPROVED_RUN, _COMPONENTS))
    # Summary block includes the quarter (row 7 = 5th summary row).
    labels = [ws.cell(row=r, column=1).value for r in range(3, 10)]
    assert "Quarter" in labels
    q_row = 3 + labels.index("Quarter")
    assert ws.cell(row=q_row, column=2).value == "Q2"
    # Header row: submission order — amount, avg index, base index, weight, PVC.
    header_row = 11  # title(1) + blank + 7 summary rows + blank
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, 7)]
    assert headers == [
        "Category",
        'Eligible amount "W"',
        "Average index of quarter",
        "Base index",
        "Component weight",
        "PVC amount",
    ]


def test_build_run_excel_numeric_cells_and_formats():
    ws = _load_sheet(build_run_excel(_APPROVED_RUN, _COMPONENTS))
    header_row = 11
    first = header_row + 1
    # Native numbers, not strings.
    assert float(ws.cell(row=first, column=2).value) == 100000.00
    assert float(ws.cell(row=first, column=3).value) == 133.5  # avg before base
    assert float(ws.cell(row=first, column=4).value) == 130.2
    assert float(ws.cell(row=first, column=5).value) == 0.10
    assert float(ws.cell(row=first, column=6).value) == 253.45
    # Submission-style number formats.
    assert ws.cell(row=first, column=2).number_format == "#,##0.00"
    assert ws.cell(row=first, column=3).number_format == "0.00"
    assert ws.cell(row=first, column=5).number_format == "0%"
    assert ws.cell(row=first, column=6).number_format == "#,##0.00"


def test_build_run_excel_total_is_live_sum_formula():
    ws = _load_sheet(build_run_excel(_APPROVED_RUN, _COMPONENTS))
    total_row = 11 + 1 + len(_COMPONENTS)
    assert ws.cell(row=total_row, column=1).value == "Total PVC"
    assert ws.cell(row=total_row, column=6).value == "=SUM(F12:F13)"
    assert ws.cell(row=total_row, column=6).number_format == "#,##0.00"


def test_build_run_excel_empty_components_total_is_zero():
    ws = _load_sheet(build_run_excel(_APPROVED_RUN, []))
    total_row = 11 + 1  # header row + total row directly after
    assert ws.cell(row=total_row, column=1).value == "Total PVC"
    assert float(ws.cell(row=total_row, column=6).value) == 0.0


# ── Phase 8: multi-sheet audit workbook ─────────────────────────────────────


def _build_phase8_workbook():
    return _load_workbook(
        build_run_excel(
            _APPROVED_RUN,
            _COMPONENTS,
            contract=_APPROVED_RUN,
            all_runs=_APPROVED_RUNS,
        )
    )


def test_build_run_excel_has_phase8_sheet_order():
    wb = _build_phase8_workbook()
    assert wb.sheetnames == ["Cover", "Bill", "W Derivation"]


def test_build_run_excel_cover_has_contract_fields_and_run_summary():
    cover = _build_phase8_workbook()["Cover"]
    assert cover["B3"].value == "T-123 — Bridge renewal"
    assert cover["B4"].value == "LOA-42"
    assert cover["B5"].value.date() == date(2025, 1, 15)
    assert cover["B6"].value == "Acme Infra"
    assert cover["B7"].value.date() == date(2025, 1, 1)
    assert cover["B8"].value == "WR"
    assert [cover.cell(row=11, column=c).value for c in range(1, 5)] == [
        "Bill No.",
        "Quarter",
        "W Amount",
        "Total PVC",
    ]
    assert [cover.cell(row=12, column=c).value for c in range(1, 5)] == [
        "RA-2",
        "Q2",
        425000,
        3801.55,
    ]
    assert cover["A12"].font.bold is True
    assert cover["A15"].value == "Generated by TenderAudit"


def test_build_run_excel_w_derivation_is_auditable():
    ws = _build_phase8_workbook()["W Derivation"]
    assert ws["A4"].value == "On-account (gross) amount"
    assert ws["B4"].value == 500000
    assert ws["A10"].value == "Less: technical withheld"
    assert ws["B10"].value == 7500
    assert ws["A11"].value == "Less: recoveries affecting PVC"
    assert ws["B11"].value == 10000
    assert ws["A13"].value == "W (eligible amount)"
    assert ws["B13"].value == 425000
    assert ws["A15"].value == "Prior negative PVC carry-forward (affects total PVC)"
    assert ws["B15"].value == 1200

"""Render an approved PVC run to a downloadable Excel workbook or PDF.

These are pure functions: they take the already-fetched run + component rows
(the same shapes `GET /api/pvc-runs/{id}` returns) and return file bytes. The
route layer owns tenant/status gating; this module owns formatting only.

Library choice (SH-P5-5/6): `openpyxl` (Excel) and `fpdf2` (PDF) are both
pure-Python with no native/system dependencies, so a clean checkout boots
from declared deps alone — a hard requirement on the Windows dev/test env
(WeasyPrint's GTK/Pango/Cairo stack is not pip-installable there).

The Excel export is a three-sheet audit workbook: contract/run history on
Cover, the existing submission-parity component table on Bill, and the
persisted W arithmetic on W Derivation. Steel remains one component row until
the engine exposes cost-driver sub-lines.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Any

# Component table columns, in submission-sheet display order. Keys match the
# `pvc_components` rows returned by GET /api/pvc-runs/{id}. The third field
# is the Excel number format ("" = text/label column).
_FMT_MONEY = "#,##0.00"
_FMT_INDEX = "0.00"
_FMT_WEIGHT = "0%"

_COMPONENT_COLUMNS: list[tuple[str, str, str]] = [
    ("category", "Category", ""),
    ("eligible_amount", 'Eligible amount "W"', _FMT_MONEY),
    ("current_avg_index", "Average index of quarter", _FMT_INDEX),
    ("base_index", "Base index", _FMT_INDEX),
    ("weight", "Component weight", _FMT_WEIGHT),
    ("pvc_value", "PVC amount", _FMT_MONEY),
]


def _fmt(value: Any) -> str:
    """Render a cell value as a plain string, preserving Decimal precision."""
    if value is None:
        return ""
    return str(value)


def _num(value: Any) -> Decimal | None:
    """Coerce a numeric cell value to Decimal (None passes through)."""
    if value is None:
        return None
    return Decimal(str(value))


def _total_pvc(components: list[dict[str, Any]]) -> Decimal:
    total = Decimal("0")
    for c in components:
        raw = c.get("pvc_value")
        if raw is not None:
            total += Decimal(str(raw))
    return total


def _summary_rows(run: dict[str, Any]) -> list[tuple[str, str]]:
    return [
        ("Run ID", _fmt(run.get("id"))),
        ("Status", _fmt(run.get("status"))),
        ("Tender", _fmt(run.get("tender_number"))),
        ("Contractor", _fmt(run.get("contractor_name"))),
        ("Quarter", _fmt(run.get("quarter_used"))),
        ("Approved by", _fmt(run.get("approved_by"))),
        ("Approved at", _fmt(run.get("approved_at"))),
    ]


def _configure_print_sheet(ws: Any, *, landscape: bool = False) -> None:
    """Keep each audit sheet readable when printed or converted to PDF."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3


def _build_cover_sheet(
    ws: Any,
    run: dict[str, Any],
    contract: dict[str, Any],
    all_runs: list[dict[str, Any]],
) -> None:
    from openpyxl.styles import Font, PatternFill

    bold = Font(bold=True)
    highlight = PatternFill(fill_type="solid", fgColor="FFF2CC")

    ws["A1"] = "TenderAudit PVC Submission"
    ws["A1"].font = Font(bold=True, size=14)

    tender = _fmt(contract.get("tender_number"))
    description = _fmt(contract.get("work_description"))
    work = f"{tender} — {description}" if tender and description else tender or description
    reference_rows = [
        ("Work", work),
        ("LOA No.", contract.get("loa_number")),
        ("LOA Date", contract.get("loa_date")),
        ("Contractor", contract.get("contractor_name")),
        ("Base Month", contract.get("base_month")),
        ("Railway Zone", contract.get("railway_zone")),
    ]
    for row, (label, value) in enumerate(reference_rows, start=3):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)

    ws["A10"] = "Approved PVC Runs"
    ws["A10"].font = bold
    headers = ("Bill No.", "Quarter", "W Amount", "Total PVC")
    for col, header in enumerate(headers, start=1):
        ws.cell(row=11, column=col, value=header).font = bold

    for row, sibling in enumerate(all_runs, start=12):
        values = (
            sibling.get("bill_number"),
            sibling.get("quarter_used"),
            _num(sibling.get("w_amount")),
            _num(sibling.get("total_pvc")),
        )
        is_current_run = sibling.get("id") == run.get("id")
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            if col >= 3:
                cell.number_format = _FMT_MONEY
            if is_current_run:
                cell.font = bold
                cell.fill = highlight

    note_row = 13 + len(all_runs)
    ws.cell(row=note_row, column=1, value="Generated by TenderAudit")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    _configure_print_sheet(ws, landscape=True)


def _build_bill_sheet(
    ws: Any,
    run: dict[str, Any],
    components: list[dict[str, Any]],
) -> None:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)

    ws["A1"] = "PVC Run Export"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    for label, value in _summary_rows(run):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    header_row = row
    for col, (_, header, _numfmt) in enumerate(_COMPONENT_COLUMNS, start=1):
        ws.cell(row=header_row, column=col, value=header).font = bold
    row += 1

    first_data_row = row
    for comp in components:
        # TODO P8-STEEL: sub-lines pending engine granularity
        for col, (key, _, numfmt) in enumerate(_COMPONENT_COLUMNS, start=1):
            if numfmt:
                cell = ws.cell(row=row, column=col, value=_num(comp.get(key)))
                cell.number_format = numfmt
            else:
                ws.cell(row=row, column=col, value=_fmt(comp.get(key)))
        row += 1

    total_col = len(_COMPONENT_COLUMNS)
    ws.cell(row=row, column=1, value="Total PVC").font = bold
    if components:
        letter = get_column_letter(total_col)
        total_cell = ws.cell(
            row=row,
            column=total_col,
            value=f"=SUM({letter}{first_data_row}:{letter}{row - 1})",
        )
    else:
        total_cell = ws.cell(row=row, column=total_col, value=_num(Decimal("0")))
    total_cell.font = bold
    total_cell.number_format = _FMT_MONEY
    widths = (20, 22, 25, 18, 20, 18)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    _configure_print_sheet(ws, landscape=True)


def _build_w_derivation_sheet(ws: Any, run: dict[str, Any]) -> None:
    from openpyxl.styles import Font

    bold = Font(bold=True)
    derivation = run.get("w_derivation")
    if not isinstance(derivation, dict):
        derivation = {}
    bill_snapshot = run.get("bill_snapshot")
    if not isinstance(bill_snapshot, dict):
        bill_snapshot = {}

    ws["A1"] = "W Derivation"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Step"
    ws["B3"] = "Amount"
    ws["A3"].font = bold
    ws["B3"].font = bold

    rows = [
        ("On-account (gross) amount", "on_account_amount"),
        ("Less: cement", "cement"),
        ("Less: steel — angles", "steel_angles"),
        ("Less: steel — plates", "steel_plates"),
        ("Less: steel — TMT", "steel_tmt"),
        ("Less: steel — other", "steel_other"),
        ("Less: technical withheld", "technical_withheld"),
        ("Less: recoveries affecting PVC", "recoveries_affecting_pvc"),
        ("Less: excluded extra items", "extra_items"),
        ("W (eligible amount)", "w"),
    ]
    for row, (label, key) in enumerate(rows, start=4):
        ws.cell(row=row, column=1, value=label)
        amount = ws.cell(row=row, column=2, value=_num(derivation.get(key)))
        amount.number_format = _FMT_MONEY
        if key == "w":
            ws.cell(row=row, column=1).font = bold
            amount.font = bold

    ws["A15"] = "Prior negative PVC carry-forward (affects total PVC)"
    ws["B15"] = _num(bill_snapshot.get("prior_negative_carry_forward"))
    ws["B15"].number_format = _FMT_MONEY
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 18
    _configure_print_sheet(ws)


def build_run_excel(
    run: dict[str, Any],
    components: list[dict[str, Any]],
    contract: dict[str, Any] | None = None,
    all_runs: list[dict[str, Any]] | None = None,
) -> bytes:
    """Return the Phase 8 three-sheet .xlsx audit workbook as bytes."""
    from openpyxl import Workbook

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    bill = wb.create_sheet("Bill")
    w_derivation = wb.create_sheet("W Derivation")

    _build_cover_sheet(cover, run, contract or {}, all_runs or [])
    _build_bill_sheet(bill, run, components)
    _build_w_derivation_sheet(w_derivation, run)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_cell_text(key: str, value: Any) -> str:
    """Format a component value for the PDF table (mirrors the Excel formats)."""
    if value is None:
        return ""
    if key == "weight":
        return f"{Decimal(str(value)) * 100:.0f}%"
    if key in ("eligible_amount", "pvc_value"):
        return f"{Decimal(str(value)):,.2f}"
    if key in ("base_index", "current_avg_index"):
        return f"{Decimal(str(value)):.2f}"
    return str(value)


def build_run_pdf(run: dict[str, Any], components: list[dict[str, Any]]) -> bytes:
    """Return a PDF report for the run as bytes."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "PVC Run Export", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    pdf.set_font("Helvetica", size=10)
    for label, value in _summary_rows(run):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(45, 6, f"{label}:")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, value, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    col_w = (pdf.w - pdf.l_margin - pdf.r_margin) / len(_COMPONENT_COLUMNS)
    pdf.set_font("Helvetica", "B", 9)
    for _, header, _numfmt in _COMPONENT_COLUMNS:
        pdf.cell(col_w, 7, header, border=1)
    pdf.ln()

    pdf.set_font("Helvetica", size=9)
    for comp in components:
        for key, _, _numfmt in _COMPONENT_COLUMNS:
            pdf.cell(col_w, 6, _pdf_cell_text(key, comp.get(key)), border=1)
        pdf.ln()

    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(col_w * (len(_COMPONENT_COLUMNS) - 1), 7, "Total PVC", border=1)
    pdf.cell(col_w, 7, f"{_total_pvc(components):,.2f}", border=1)

    out = pdf.output()
    return bytes(out)
